import os, json, logging, sys, zipfile, tempfile, shutil
from datetime import datetime, timedelta
from flask import Flask, request, jsonify, send_from_directory, send_file, redirect, url_for
from flask_cors import CORS
import hashlib, uuid
from werkzeug.utils import secure_filename
from order_manager import order_manager
import database as db
import openpyxl

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
print("=" * 60)
print("MA FURNITURE - SQLITE STORAGE")
print("=" * 60)
print(f"Python version: {sys.version}")
print(f"Current working directory: {os.getcwd()}")
print(f"BASE_DIR: {BASE_DIR}")

DATA_DIR = os.path.join(BASE_DIR, 'data')
STATIC_DIR = os.path.join(BASE_DIR, 'static')
UPLOAD_FOLDER = os.path.join(STATIC_DIR, 'uploads/products')
TEMP_FOLDER = os.path.join(STATIC_DIR, 'uploads/temp')
DB_PATH = os.path.join(DATA_DIR, 'ma_furniture.db')
PRODUCTS_DIR = os.path.join(DATA_DIR, 'products')
SECTIONS_FILE = os.path.join(DATA_DIR, 'sections.json')
BACKGROUND_FILE = os.path.join(DATA_DIR, 'background.json')

os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(TEMP_FOLDER, exist_ok=True)

app = Flask(__name__, static_folder='static', static_url_path='')
CORS(app)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['TEMP_FOLDER'] = TEMP_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 200 * 1024 * 1024
app.config['UPLOAD_TIMEOUT'] = 300
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}

logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = app.logger

db.init_db()
db.migrate_from_json(PRODUCTS_DIR, SECTIONS_FILE, BACKGROUND_FILE, DATA_DIR)

# Вспомогательные функции (без изменений)
def allowed_file(filename): return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS
def generate_filename(original_name):
    timestamp = int(datetime.now().timestamp())
    random_str = hashlib.md5(str(uuid.uuid4()).encode()).hexdigest()[:8]
    ext = original_name.rsplit('.', 1)[1].lower() if '.' in original_name else 'jpg'
    return f"{timestamp}_{random_str}.{ext}"
def get_uploads_path(): return os.path.join(STATIC_DIR, 'uploads', 'products')

def get_all_products():
    with db.get_db() as conn:
        rows=conn.execute('SELECT id, name, code, category, section, price, old_price, badge, recommended, description, specifications, status, stock, images, color_variants, created_at, updated_at, is_price_on_request, availability FROM products ORDER BY created_at DESC').fetchall()
        products=[]
        for row in rows:
            prod=dict(row)
            prod['images']=json.loads(prod['images']) if prod['images'] else []
            prod['color_variants']=json.loads(prod['color_variants']) if prod['color_variants'] else []
            products.append(prod)
        return products

def get_product_by_id(product_id):
    with db.get_db() as conn:
        row=conn.execute('SELECT id, name, code, category, section, price, old_price, badge, recommended, description, specifications, status, stock, images, color_variants, created_at, updated_at, is_price_on_request, availability FROM products WHERE id = ?',(product_id,)).fetchone()
        if row:
            prod=dict(row)
            prod['images']=json.loads(prod['images']) if prod['images'] else []
            prod['color_variants']=json.loads(prod['color_variants']) if prod['color_variants'] else []
            return prod
        return None

def save_product(product_data):
    now=datetime.now().isoformat()
    if 'id' in product_data and product_data['id']:
        product_id=product_data['id']
        with db.get_db() as conn:
            conn.execute('''UPDATE products SET name=?, code=?, category=?, section=?, price=?, old_price=?, badge=?,
                recommended=?, description=?, specifications=?, status=?, stock=?, images=?, color_variants=?, updated_at=?, is_price_on_request=?, availability=?
                WHERE id=?''',(product_data.get('name'),product_data.get('code'),product_data.get('category'),
                product_data.get('section'),product_data.get('price',0),product_data.get('old_price'),product_data.get('badge'),
                1 if product_data.get('recommended') else 0,product_data.get('description'),product_data.get('specifications'),
                product_data.get('status','active'),product_data.get('stock',0),json.dumps(product_data.get('images',[])),
                json.dumps(product_data.get('color_variants',[])),now,1 if product_data.get('is_price_on_request') else 0,product_data.get('availability', 0),product_id))
        return product_id
    else:
        with db.get_db() as conn:
            cur=conn.execute('SELECT MAX(id) FROM products')
            max_id=cur.fetchone()[0] or 0
            new_id=max_id+1
            product_data['id']=new_id
            product_data['created_at']=now
            product_data['updated_at']=now
            conn.execute('''INSERT INTO products (id,name,code,category,section,price,old_price,badge,recommended,description,
                specifications,status,stock,images,color_variants,created_at,updated_at,is_price_on_request,availability)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',(new_id,product_data.get('name'),product_data.get('code'),
                product_data.get('category'),product_data.get('section'),product_data.get('price',0),product_data.get('old_price'),
                product_data.get('badge'),1 if product_data.get('recommended') else 0,product_data.get('description'),
                product_data.get('specifications'),product_data.get('status','active'),product_data.get('stock',0),
                json.dumps(product_data.get('images',[])),json.dumps(product_data.get('color_variants',[])),now,now,
                1 if product_data.get('is_price_on_request') else 0,product_data.get('availability', 0)))
        return new_id

def delete_product(product_id):
    with db.get_db() as conn:
        conn.execute('DELETE FROM products WHERE id = ?',(product_id,))
        return True

def load_sections():
    with db.get_db() as conn:
        rows=conn.execute('SELECT * FROM sections ORDER BY display_order').fetchall()
        sections=[dict(row) for row in rows]
        if not sections:
            default_sections=[{"id":1,"code":"pantographs","name":"Пантографы","active":True,"display_order":1},
                {"id":2,"code":"wardrobes","name":"Гардеробные системы","active":True,"display_order":2},
                {"id":3,"code":"shoeracks","name":"Обувницы","active":True,"display_order":3}]
            for s in default_sections:
                conn.execute('INSERT INTO sections (id,code,name,active,display_order) VALUES (?,?,?,?,?)',
                    (s['id'],s['code'],s['name'],1,s['display_order']))
            return default_sections
        return sections

def save_sections(sections):
    with db.get_db() as conn:
        conn.execute('DELETE FROM sections')
        for s in sections:
            conn.execute('INSERT INTO sections (id,code,name,active,display_order) VALUES (?,?,?,?,?)',
                (s.get('id'),s.get('code'),s.get('name'),1 if s.get('active') else 0,s.get('display_order',0)))

def load_background():
    with db.get_db() as conn:
        row=conn.execute('SELECT * FROM backgrounds LIMIT 1').fetchone()
        return dict(row) if row else None

def save_background(background_data):
    now=datetime.now().isoformat()
    with db.get_db() as conn:
        if 'id' in background_data and background_data['id']:
            conn.execute('UPDATE backgrounds SET title=?, description=?, image_url=?, active=?, updated_at=? WHERE id=?',
                (background_data.get('title'),background_data.get('description'),background_data.get('image_url'),
                 1 if background_data.get('active',True) else 0,now,background_data['id']))
        else:
            conn.execute('INSERT INTO backgrounds (title,description,image_url,active,created_at,updated_at) VALUES (?,?,?,?,?,?)',
                (background_data.get('title'),background_data.get('description'),background_data.get('image_url'),
                 1 if background_data.get('active',True) else 0,now,now))

def get_recent_activity():
    activity=[]
    products=get_all_products()
    recent_products=sorted(products,key=lambda x: x.get('created_at',''),reverse=True)[:3]
    for product in recent_products:
        activity.append({'type':'product','title':'Добавлен новый товар','description':product.get('name','Товар'),
            'time':product.get('created_at',''),'icon':'fas fa-box'})
    background=load_background()
    if background and background.get('updated_at'):
        activity.append({'type':'background','title':'Обновлен фон','description':background.get('title','Главный фон'),
            'time':background.get('updated_at',''),'icon':'fas fa-image'})
    activity.sort(key=lambda x: x.get('time',''),reverse=True)
    return activity

# ==================== НОВЫЕ МАРШРУТЫ (ЧПУ) ====================

@app.route('/')
def index():
    try:
        index_path = os.path.join(STATIC_DIR, 'index.html')
        if os.path.exists(index_path):
            return send_file(index_path)
        return '<!DOCTYPE html>...'  # fallback
    except Exception as e:
        return f'Error: {str(e)}', 500

@app.route('/shop')
def shop():
    try:
        shop_path = os.path.join(STATIC_DIR, 'shop.html')
        if os.path.exists(shop_path):
            return send_file(shop_path)
        return "Shop page not found", 404
    except Exception as e:
        return str(e), 500

@app.route('/product/<int:product_id>')
def product_page(product_id):
    """Страница отдельного товара по ID"""
    try:
        piece_path = os.path.join(STATIC_DIR, 'piece.html')
        if os.path.exists(piece_path):
            return send_file(piece_path)
        return "Product page not found", 404
    except Exception as e:
        return str(e), 500

# Для обратной совместимости: /piece?id=...
@app.route('/piece')
def piece_legacy():
    product_id = request.args.get('id')
    if product_id:
        return redirect(url_for('product_page', product_id=product_id), code=301)
    # Если нет id, просто показываем страницу (хотя она без товара бесполезна)
    piece_path = os.path.join(STATIC_DIR, 'piece.html')
    if os.path.exists(piece_path):
        return send_file(piece_path)
    return "Product page not found", 404

@app.route('/order-success')
def order_success():
    try:
        success_path = os.path.join(STATIC_DIR, 'order-success.html')
        if os.path.exists(success_path):
            return send_file(success_path)
        return '<!DOCTYPE html>...', 200
    except Exception as e:
        return str(e), 500

# Админка
@app.route('/admin')
@app.route('/admin/')
def admin_login():
    try:
        login_path = os.path.join(STATIC_DIR, 'admin', 'admin-login.html')
        if os.path.exists(login_path):
            return send_file(login_path)
        return "Admin login not found", 404
    except Exception as e:
        return str(e), 500

@app.route('/admin/dashboard')
@app.route('/admin/dashboard/')
def admin_dashboard():
    try:
        dashboard_path = os.path.join(STATIC_DIR, 'admin', 'admin-dashboard.html')
        if os.path.exists(dashboard_path):
            return send_file(dashboard_path)
        return "Dashboard not found", 404
    except Exception as e:
        return str(e), 500

@app.route('/admin/products')
@app.route('/admin/products/')
def admin_products():
    try:
        products_path = os.path.join(STATIC_DIR, 'admin', 'products-management.html')
        if os.path.exists(products_path):
            return send_file(products_path)
        return "Products management not found", 404
    except Exception as e:
        return str(e), 500

@app.route('/admin/categories')
@app.route('/admin/categories/')
def admin_categories():
    try:
        categories_path = os.path.join(STATIC_DIR, 'admin', 'categories-management.html')
        if os.path.exists(categories_path):
            return send_file(categories_path)
        return "Categories management not found", 404
    except Exception as e:
        return str(e), 500

@app.route('/admin/media')
@app.route('/admin/media/')
def admin_media():
    try:
        media_path = os.path.join(STATIC_DIR, 'admin', 'media-management.html')
        if os.path.exists(media_path):
            return send_file(media_path)
        return "Media management not found", 404
    except Exception as e:
        return str(e), 500

@app.route('/admin/backup')
def admin_backup():
    try:
        backup_path = os.path.join(STATIC_DIR, 'admin', 'backup-management.html')
        if os.path.exists(backup_path):
            return send_file(backup_path)
        return "Backup page not found", 404
    except Exception as e:
        return str(e), 500

@app.route('/admin/techdocks')
def admin_techdocks():
    try:
        tech_path = os.path.join(STATIC_DIR, 'admin', 'techdocks.html')
        if os.path.exists(tech_path):
            return send_file(tech_path)
        return "Tech docs not found", 404
    except Exception as e:
        return str(e), 500

# Редиректы со старых .html URL на новые (301 Moved Permanently)
@app.route('/index.html')
def redirect_index():
    return redirect('/', code=301)

@app.route('/shop.html')
def redirect_shop():
    return redirect('/shop', code=301)

@app.route('/piece.html')
def redirect_piece():
    # Если есть id, перенаправляем на /product/<id>
    product_id = request.args.get('id')
    if product_id:
        return redirect(url_for('product_page', product_id=product_id), code=301)
    return redirect('/piece', code=301)

@app.route('/order-success.html')
def redirect_order_success():
    return redirect('/order-success', code=301)

@app.route('/admin.html')
def redirect_admin():
    return redirect('/admin', code=301)

# ========== НОВЫЕ ЭНДПОИНТЫ ДЛЯ ОНЛАЙН-ТРЕКЕРА ==========
active_sessions={}
@app.route('/api/admin/online/count',methods=['GET'])
def admin_online_count():
    try:
        now=datetime.now()
        cutoff=now-timedelta(minutes=5)
        expired=[ip for ip,ts in active_sessions.items() if ts<cutoff]
        for ip in expired: del active_sessions[ip]
        client_ip=request.remote_addr
        active_sessions[client_ip]=now
        return jsonify({'success':True,'count':len(active_sessions)})
    except Exception as e: return jsonify({'success':False,'error':str(e)}),500

# ========== РАСШИРЕННЫЕ ЭНДПОИНТЫ БЭКАПОВ ==========
@app.route('/api/admin/backup/download',methods=['GET'])
def admin_backup_download():
    try:
        token=request.headers.get('Authorization')
        if not token: return jsonify({'success':False,'error':'Требуется авторизация'}),401
        temp_dir=tempfile.mkdtemp()
        try:
            backup_db_path=os.path.join(temp_dir,'ma_furniture.db')
            shutil.copy2(DB_PATH,backup_db_path)
            uploads_src=get_uploads_path()
            uploads_dst=os.path.join(temp_dir,'uploads','products')
            if os.path.exists(uploads_src) and os.listdir(uploads_src): shutil.copytree(uploads_src,uploads_dst)
            else: os.makedirs(uploads_dst,exist_ok=True)
            zip_path=os.path.join(temp_dir,'ma_furniture_backup.zip')
            with zipfile.ZipFile(zip_path,'w',zipfile.ZIP_DEFLATED) as zf:
                zf.write(backup_db_path,arcname='ma_furniture.db')
                for root,dirs,files in os.walk(uploads_dst):
                    for file in files:
                        full_path=os.path.join(root,file)
                        rel_path=os.path.relpath(full_path,temp_dir)
                        zf.write(full_path,arcname=rel_path)
            backup_dir=os.path.join(DATA_DIR,'backups')
            os.makedirs(backup_dir,exist_ok=True)
            backup_copy_path=os.path.join(backup_dir,f'ma_furniture_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.zip')
            shutil.copy2(zip_path,backup_copy_path)
            return send_file(zip_path,as_attachment=True,
                download_name=f'ma_furniture_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.zip',mimetype='application/zip')
        finally: shutil.rmtree(temp_dir,ignore_errors=True)
    except Exception as e: return jsonify({'success':False,'error':str(e)}),500

@app.route('/api/admin/backup/upload',methods=['POST'])
def admin_backup_upload():
    try:
        token=request.headers.get('Authorization')
        if not token: return jsonify({'success':False,'error':'Требуется авторизация'}),401
        if 'file' not in request.files: return jsonify({'success':False,'error':'Файл не загружен'}),400
        file=request.files['file']
        if file.filename=='': return jsonify({'success':False,'error':'Файл не выбран'}),400
        if not file.filename.endswith('.zip'): return jsonify({'success':False,'error':'Допустим только ZIP-архив'}),400
        temp_dir=tempfile.mkdtemp()
        try:
            zip_path=os.path.join(temp_dir,'upload.zip')
            file.save(zip_path)
            with zipfile.ZipFile(zip_path,'r') as zf:
                if 'ma_furniture.db' not in zf.namelist(): return jsonify({'success':False,'error':'Архив не содержит файл ma_furniture.db'}),400
                zf.extractall(temp_dir)
            extracted_db=os.path.join(temp_dir,'ma_furniture.db')
            if not os.path.exists(extracted_db): return jsonify({'success':False,'error':'Не удалось извлечь базу данных'}),500
            extracted_uploads=os.path.join(temp_dir,'uploads','products')
            has_uploads=os.path.exists(extracted_uploads) and os.listdir(extracted_uploads)
            backup_db_path=DB_PATH+'.backup'
            if os.path.exists(DB_PATH): shutil.copy2(DB_PATH,backup_db_path)
            uploads_path=get_uploads_path()
            backup_uploads_path=uploads_path+'.backup'
            if os.path.exists(uploads_path):
                if os.path.exists(backup_uploads_path): shutil.rmtree(backup_uploads_path)
                shutil.copytree(uploads_path,backup_uploads_path)
            shutil.copy2(extracted_db,DB_PATH)
            if has_uploads:
                if os.path.exists(uploads_path): shutil.rmtree(uploads_path)
                shutil.copytree(extracted_uploads,uploads_path)
            else: logger.info("Архив не содержит изображений, папка uploads не изменена")
            return jsonify({'success':True,'message':'База данных и изображения успешно восстановлены. Перезагрузите страницу.'})
        except Exception as e:
            logger.error(f"Ошибка восстановления из бэкапа: {e}")
            try:
                if os.path.exists(DB_PATH+'.backup'): shutil.copy2(DB_PATH+'.backup',DB_PATH)
                uploads_path=get_uploads_path()
                if os.path.exists(uploads_path+'.backup'):
                    if os.path.exists(uploads_path): shutil.rmtree(uploads_path)
                    shutil.copytree(uploads_path+'.backup',uploads_path)
            except: pass
            return jsonify({'success':False,'error':str(e)}),500
        finally: shutil.rmtree(temp_dir,ignore_errors=True)
    except Exception as e: return jsonify({'success':False,'error':str(e)}),500

@app.route('/api/admin/backup/list',methods=['GET'])
def admin_backup_list():
    try:
        token=request.headers.get('Authorization')
        if not token: return jsonify({'success':False,'error':'Требуется авторизация'}),401
        backup_dir=os.path.join(DATA_DIR,'backups')
        os.makedirs(backup_dir,exist_ok=True)
        backups=[]
        for filename in os.listdir(backup_dir):
            if filename.endswith('.zip'):
                filepath=os.path.join(backup_dir,filename)
                stat=os.stat(filepath)
                backups.append({'filename':filename,'size':stat.st_size,'created':datetime.fromtimestamp(stat.st_ctime).isoformat()})
        backups.sort(key=lambda x: x['created'],reverse=True)
        return jsonify({'success':True,'backups':backups})
    except Exception as e: return jsonify({'success':False,'error':str(e)}),500

@app.route('/api/admin/backup/delete',methods=['POST'])
def admin_backup_delete():
    try:
        token=request.headers.get('Authorization')
        if not token: return jsonify({'success':False,'error':'Требуется авторизация'}),401
        data=request.get_json()
        filename=data.get('filename')
        if not filename: return jsonify({'success':False,'error':'Не указано имя файла'}),400
        backup_dir=os.path.join(DATA_DIR,'backups')
        filepath=os.path.join(backup_dir,filename)
        if os.path.exists(filepath):
            os.remove(filepath)
            return jsonify({'success':True,'message':'Бэкап удалён'})
        else: return jsonify({'success':False,'error':'Файл не найден'}),404
    except Exception as e: return jsonify({'success':False,'error':str(e)}),500

# ========== ОСТАЛЬНЫЕ API ==========
@app.route('/api/admin/sections',methods=['GET'])
def admin_get_sections():
    try:
        sections=load_sections()
        return jsonify({'success':True,'sections':sections})
    except Exception as e: return jsonify({'success':False,'error':str(e)}),500

@app.route('/api/admin/sections',methods=['POST'])
def admin_create_section():
    try:
        data=request.get_json()
        required_fields=['name','code']
        for field in required_fields:
            if field not in data or not str(data[field]).strip():
                return jsonify({'success':False,'error':f'Поле {field} обязательно'}),400
        sections=load_sections()
        new_id=max([s.get('id',0) for s in sections],default=0)+1
        if any(s.get('code')==data['code'] for s in sections):
            return jsonify({'success':False,'error':'Раздел с таким кодом уже существует'}),400
        new_section={'id':new_id,'name':data['name'].strip(),'code':data['code'].strip().lower(),
            'active':data.get('active',True),'display_order':data.get('display_order',len(sections)+1)}
        sections.append(new_section)
        save_sections(sections)
        return jsonify({'success':True,'section':new_section,'message':'Раздел успешно создан'})
    except Exception as e: return jsonify({'success':False,'error':str(e)}),500

@app.route('/api/admin/sections/<int:section_id>',methods=['PUT'])
def admin_update_section(section_id):
    try:
        data=request.get_json()
        sections=load_sections()
        section_index=None
        for i,section in enumerate(sections):
            if section.get('id')==section_id: section_index=i; break
        if section_index is None: return jsonify({'success':False,'error':'Раздел не найден'}),404
        if 'name' in data: sections[section_index]['name']=data['name'].strip()
        if 'code' in data:
            new_code=data['code'].strip().lower()
            if any(s.get('code')==new_code and s.get('id')!=section_id for s in sections):
                return jsonify({'success':False,'error':'Раздел с таким кодом уже существует'}),400
            sections[section_index]['code']=new_code
        if 'active' in data: sections[section_index]['active']=bool(data['active'])
        if 'display_order' in data: sections[section_index]['display_order']=int(data['display_order'])
        save_sections(sections)
        return jsonify({'success':True,'section':sections[section_index],'message':'Раздел успешно обновлен'})
    except Exception as e: return jsonify({'success':False,'error':str(e)}),500

@app.route('/api/admin/sections/<int:section_id>',methods=['DELETE'])
def admin_delete_section(section_id):
    try:
        sections=load_sections()
        section_index=None
        section_to_delete=None
        for i,section in enumerate(sections):
            if section.get('id')==section_id: section_index=i; section_to_delete=section; break
        if section_index is None: return jsonify({'success':False,'error':'Раздел не найден'}),404
        products=get_all_products()
        products_in_section=[p for p in products if p.get('section')==section_to_delete.get('code')]
        if products_in_section:
            return jsonify({'success':False,'error':f'Нельзя удалить раздел, так как в нем есть товары ({len(products_in_section)} шт.)'}),400
        sections.pop(section_index)
        save_sections(sections)
        return jsonify({'success':True,'message':'Раздел успешно удален'})
    except Exception as e: return jsonify({'success':False,'error':str(e)}),500

@app.route('/api/admin/sections/reorder',methods=['POST'])
def admin_reorder_sections():
    try:
        data=request.get_json()
        new_order=data.get('order',[])
        if not new_order: return jsonify({'success':False,'error':'Не указан новый порядок'}),400
        sections=load_sections()
        for section in sections:
            if str(section.get('id')) in new_order:
                section['display_order']=new_order.index(str(section.get('id')))+1
        save_sections(sections)
        return jsonify({'success':True,'message':'Порядок разделов обновлен'})
    except Exception as e: return jsonify({'success':False,'error':str(e)}),500

@app.route('/api/media/background',methods=['GET'])
def get_background():
    try:
        background=load_background()
        if not background: return jsonify({'success':False,'error':'Фон не найден','background':None}),404
        return jsonify({'success':True,'background':background})
    except Exception as e: return jsonify({'success':False,'error':str(e)}),500

@app.route('/api/admin/media/background',methods=['GET'])
def admin_get_background():
    try:
        background=load_background()
        return jsonify({'success':True,'background':background})
    except Exception as e: return jsonify({'success':False,'error':str(e)}),500

@app.route('/api/admin/media/background',methods=['POST'])
def admin_create_background():
    try:
        data=request.get_json()
        if not data.get('image_url'): return jsonify({'success':False,'error':'Изображение обязательно'}),400
        if save_background(data):
            background=load_background()
            return jsonify({'success':True,'background':background,'message':'Фон успешно создан'})
        else: return jsonify({'success':False,'error':'Ошибка сохранения'}),500
    except Exception as e: return jsonify({'success':False,'error':str(e)}),500

@app.route('/api/admin/media/background/<int:background_id>',methods=['PUT'])
def admin_update_background(background_id):
    try:
        data=request.get_json()
        current_background=load_background()
        if not current_background or current_background.get('id')!=background_id:
            return jsonify({'success':False,'error':'Фон не найден'}),404
        for key,value in data.items():
            if key in ['title','description','image_url','active']: current_background[key]=value
        current_background['updated_at']=datetime.now().isoformat()
        if save_background(current_background):
            return jsonify({'success':True,'background':current_background,'message':'Фон успешно обновлен'})
        else: return jsonify({'success':False,'error':'Ошибка сохранения'}),500
    except Exception as e: return jsonify({'success':False,'error':str(e)}),500

@app.route('/api/admin/media/background/<int:background_id>',methods=['DELETE'])
def admin_delete_background(background_id):
    try:
        current_background=load_background()
        if not current_background or current_background.get('id')!=background_id:
            return jsonify({'success':False,'error':'Фон не найден'}),404
        if os.path.exists(BACKGROUND_FILE):
            os.remove(BACKGROUND_FILE)
            image_url=current_background.get('image_url','')
            if image_url and '/uploads/' in image_url:
                try:
                    filename=image_url.split('/')[-1]
                    filepath=os.path.join(UPLOAD_FOLDER,filename)
                    if os.path.exists(filepath): os.remove(filepath)
                except: pass
            return jsonify({'success':True,'message':'Фон успешно удален'})
        else: return jsonify({'success':False,'error':'Файл фона не найден'}),404
    except Exception as e: return jsonify({'success':False,'error':str(e)}),500

@app.route('/api/admin/dashboard/stats',methods=['GET'])
def admin_dashboard_stats():
    try:
        products=get_all_products()
        active_products=len([p for p in products if p.get('status')=='active'])
        sections=load_sections()
        active_sections=len([s for s in sections if s.get('active',True)])
        background=load_background()
        background_exists=bool(background and background.get('image_url'))
        return jsonify({'success':True,'stats':{'total_products':len(products),'active_products':active_products,
            'total_sections':len(sections),'active_sections':active_sections,'background_exists':background_exists}})
    except Exception as e: return jsonify({'success':False,'error':str(e)}),500

@app.route('/api/admin/dashboard/activity',methods=['GET'])
def admin_dashboard_activity():
    try:
        activity=get_recent_activity()
        return jsonify({'success':True,'activity':activity[:5]})
    except Exception as e: return jsonify({'success':False,'error':str(e)}),500

@app.route('/api/admin/dashboard/popular-products',methods=['GET'])
def admin_dashboard_popular_products():
    try:
        products=get_all_products()
        popular_products=[]
        for product in products:
            if product.get('status')=='active':
                badge=product.get('badge','')
                if badge in ['Хит продаж','Новинка','Акция'] or product.get('recommended'):
                    popular_products.append(product)
        popular_products=popular_products[:3]
        return jsonify({'success':True,'products':popular_products})
    except Exception as e: return jsonify({'success':False,'error':str(e)}),500

@app.route('/api/admin/products',methods=['GET'])
def admin_get_products():
    try:
        products=get_all_products()
        return jsonify({'success':True,'products':products})
    except Exception as e: return jsonify({'success':False,'error':str(e)}),500

@app.route('/api/admin/products',methods=['POST'])
def admin_create_product():
    try:
        data=request.get_json()
        required_fields=['name','price','description']
        for field in required_fields:
            if field not in data or not str(data[field]).strip():
                return jsonify({'success':False,'error':f'Поле {field} обязательно'}),400
        product_id=save_product(data)
        if not product_id: return jsonify({'success':False,'error':'Ошибка сохранения'}),500
        return jsonify({'success':True,'product_id':product_id,'message':'Товар успешно создан'})
    except Exception as e: return jsonify({'success':False,'error':str(e)}),500

@app.route('/api/admin/products/<int:product_id>',methods=['PUT'])
def admin_update_product(product_id):
    try:
        data=request.get_json()
        existing=get_product_by_id(product_id)
        if not existing: return jsonify({'success':False,'error':'Товар не найден'}),404
        for key,value in data.items(): existing[key]=value
        existing['updated_at']=datetime.now().isoformat()
        save_product(existing)
        return jsonify({'success':True,'message':'Товар успешно обновлен'})
    except Exception as e: return jsonify({'success':False,'error':str(e)}),500

@app.route('/api/admin/products/<int:product_id>',methods=['DELETE'])
def admin_delete_product(product_id):
    try:
        if delete_product(product_id): return jsonify({'success':True,'message':'Товар успешно удален'})
        else: return jsonify({'success':False,'error':'Товар не найден'}),404
    except Exception as e: return jsonify({'success':False,'error':str(e)}),500

@app.route('/api/admin/colors/palette',methods=['GET'])
def get_color_palette():
    try:
        palette=[{"name":"Черный матовый","hex":"#2C2C2C"},{"name":"Белый глянцевый","hex":"#FFFFFF"},
            {"name":"Серый металлик","hex":"#7D7D7D"},{"name":"Коричневый","hex":"#8B4513"},
            {"name":"Бежевый","hex":"#F5DEB3"},{"name":"Серый бетон","hex":"#9E9E9E"},
            {"name":"Черный глянец","hex":"#1A1A1A"},{"name":"Белый матовый","hex":"#F8F8F8"}]
        return jsonify({'success':True,'palette':palette})
    except Exception as e: return jsonify({'success':False,'error':str(e)}),500

@app.route('/api/admin/login',methods=['POST'])
def admin_login():
    try:
        data=request.get_json()
        username=data.get('username','').strip()
        password=data.get('password','')
        if not username or not password: return jsonify({'success':False,'error':'Заполните все поля'}),400
        expected_username=os.environ.get("STAD")
        expected_password=os.environ.get("SUTT")
        if not expected_username or not expected_password:
            expected_username="obratites"
            expected_password="koperatoru"
            logger.warning("Используются учетные данные по умолчанию (секреты не настроены)")
        if username==expected_username and password==expected_password:
            return jsonify({'success':True,'admin':{'username':username,'role':'admin'},'session_token':str(uuid.uuid4()),'message':'Авторизация успешна'})
        else: return jsonify({'success':False,'error':'Неверный логин или пароль'}),401
    except Exception as e: return jsonify({'success':False,'error':str(e)}),500

@app.route('/api/admin/verify')
def verify_admin():
    token=request.headers.get('Authorization')
    if token: return jsonify({'success':True})
    return jsonify({'success':False,'error':'Требуется авторизация'}),401

@app.route('/api/upload',methods=['POST'])
def upload_file():
    try:
        if 'file' not in request.files: return jsonify({'success':False,'error':'Файл не загружен'}),400
        file=request.files['file']
        if file.filename=='': return jsonify({'success':False,'error':'Файл не выбран'}),400
        if not allowed_file(file.filename): return jsonify({'success':False,'error':'Недопустимый формат файла'}),400
        original_name=secure_filename(file.filename)
        filename=generate_filename(original_name)
        file_path=os.path.join(app.config['UPLOAD_FOLDER'],filename)
        file.save(file_path)
        file_url=f"/static/uploads/products/{filename}"
        logger.info(f"Файл загружен: {filename}")
        return jsonify({'success':True,'filename':filename,'original_name':original_name,'url':file_url,'size':os.path.getsize(file_path)})
    except Exception as e: return jsonify({'success':False,'error':str(e)}),500

@app.route('/api/upload/delete',methods=['POST'])
def delete_file():
    try:
        data=request.get_json()
        filename=data.get('filename')
        if not filename: return jsonify({'success':False,'error':'Имя файла не указано'}),400
        safe_filename=secure_filename(filename)
        file_path=os.path.join(app.config['UPLOAD_FOLDER'],safe_filename)
        if not os.path.exists(file_path): return jsonify({'success':False,'error':'Файл не найден'}),404
        os.remove(file_path)
        logger.info(f"Файл удален: {safe_filename}")
        return jsonify({'success':True,'message':'Файл удален'})
    except Exception as e: return jsonify({'success':False,'error':str(e)}),500

@app.route('/api/products',methods=['GET'])
def get_products():
    try:
        category=request.args.get('category','all')
        section=request.args.get('section','')
        status=request.args.get('status','active')
        search=request.args.get('search','')
        all_products=get_all_products()
        filtered_products=[]
        for product in all_products:
            if status and product.get('status')!=status: continue
            if category!='all' and product.get('category')!=category: continue
            if section and product.get('section')!=section: continue
            if search:
                search_lower=search.lower()
                name=product.get('name','').lower()
                desc=product.get('description','').lower()
                if search_lower not in name and search_lower not in desc: continue
            filtered_products.append(product)
        return jsonify({'success':True,'products':filtered_products,'total':len(filtered_products)})
    except Exception as e: return jsonify({'success':False,'error':str(e)}),500

@app.route('/api/products/<int:product_id>',methods=['GET'])
def get_product(product_id):
    try:
        product=get_product_by_id(product_id)
        if not product: return jsonify({'success':False,'error':'Товар не найден'}),404
        return jsonify({'success':True,'product':product})
    except Exception as e: return jsonify({'success':False,'error':str(e)}),500

@app.route('/api/sections',methods=['GET'])
def get_sections():
    try:
        sections=load_sections()
        return jsonify({'success':True,'sections':sections})
    except Exception as e: return jsonify({'success':False,'error':str(e)}),500

@app.route('/api/stats')
def get_stats():
    try:
        products=get_all_products()
        active_count=len([p for p in products if p.get('status')=='active'])
        return jsonify({'success':True,'products_count':len(products),'active_products':active_count})
    except Exception as e: return jsonify({'success':False,'error':str(e)}),500

@app.route('/api/products/<int:product_id>/colors',methods=['GET'])
def get_product_colors(product_id):
    try:
        product=get_product_by_id(product_id)
        if not product: return jsonify({'success':False,'error':'Товар не найден'}),404
        if 'color_variants' in product and product['color_variants']:
            variants=product['color_variants']
        else:
            variants=[{'variant_id':product.get('code',f"ID{product['id']}"),'color_name':product.get('color_name','Основной'),
                'color_hex':product.get('color_hex','#2C2C2C'),'suffix':'','price':product.get('price',0),
                'old_price':product.get('old_price'),'stock':product.get('stock',0),'images':product.get('images',[]),'is_original':True,'order':1}]
        return jsonify({'success':True,'product_id':product_id,'variants':variants,'base_name':product.get('name',''),'base_code':product.get('code','')})
    except Exception as e: return jsonify({'success':False,'error':str(e)}),500

@app.route('/api/admin/products/<int:product_id>/color-variant', methods=['POST'])
def add_color_variant(product_id):
    try:
        # Получаем товар из БД
        product = get_product_by_id(product_id)
        if not product:
            return jsonify({'success': False, 'error': 'Товар не найден'}), 404

        data = request.get_json()

        # Проверяем обязательные поля
        if 'color_name' not in data or 'color_hex' not in data:
            return jsonify({'success': False, 'error': 'Не указаны название или цвет'}), 400

        # Убеждаемся, что поле color_variants существует и является списком
        if 'color_variants' not in product or not isinstance(product['color_variants'], list):
            product['color_variants'] = []

        # Проверка лимита (максимум 5 вариантов, включая оригинал)
        if len(product['color_variants']) >= 5:
            return jsonify({'success': False, 'error': 'Максимум 5 цветовых вариантов'}), 400

        # Генерация нового variant_id
        base_code = product.get('code', f"ID{product_id}")
        new_index = len([v for v in product['color_variants'] if not v.get('is_original', False)]) + 1
        variant_id = f"{base_code}/{new_index}"

        # Формируем новый вариант
        new_variant = {
            'variant_id': variant_id,
            'color_name': data['color_name'],
            'color_hex': data['color_hex'],
            'suffix': f" - {data['color_name']}",
            'price': data.get('price', product.get('price', 0)),
            'old_price': data.get('old_price'),
            'stock': data.get('stock', 0),
            'images': data.get('images', []),
            'is_original': False,
            'order': len(product['color_variants']) + 1
        }

        # Добавляем вариант в массив
        product['color_variants'].append(new_variant)

        # Сохраняем обновлённый товар в БД (используем существующую функцию save_product)
        save_product(product)

        # Возвращаем успешный ответ
        return jsonify({
            'success': True,
            'variant': new_variant,
            'message': 'Цветовой вариант добавлен'
        })

    except Exception as e:
        logger.error(f"Ошибка добавления цветового варианта: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/health')
def health_check():
    return jsonify({'status': 'ok', 'timestamp': datetime.now().isoformat()})

@app.route('/favicon.ico')
def favicon():
    return send_from_directory(os.path.join(app.static_folder, 'images'), 'logo2.png', mimetype='image/png')

# ========== ЭКСПОРТ БАЗЫ В EXCEL ==========
@app.route('/api/admin/export/excel', methods=['GET'])
def admin_export_excel():
    """Экспорт всей базы данных в файл Excel (.xlsx)"""
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({'success': False, 'error': 'Библиотека openpyxl не установлена. Добавьте в requirements.txt'}), 500

    token = request.headers.get('Authorization')
    if not token:
        return jsonify({'success': False, 'error': 'Требуется авторизация'}), 401

    wb = Workbook()
    # Удаляем стандартный пустой лист, если есть
    if 'Sheet' in wb.sheetnames:
        std_sheet = wb['Sheet']
        wb.remove(std_sheet)

    with db.get_db() as conn:
        # Получаем список всех таблиц (кроме системных)
        tables = conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%'"
        ).fetchall()

        for table in tables:
            table_name = table['name']
            # Создаём лист с именем таблицы (обрезаем до 31 символа – ограничение Excel)
            sheet_name = table_name[:31]
            ws = wb.create_sheet(title=sheet_name)

            # Получаем данные таблицы
            rows = conn.execute(f"SELECT * FROM {table_name}").fetchall()
            if not rows:
                continue

            # Заголовки – названия колонок
            columns = rows[0].keys()
            for col_idx, col_name in enumerate(columns, 1):
                ws.cell(row=1, column=col_idx, value=col_name)

            # Данные
            for row_idx, row in enumerate(rows, 2):
                for col_idx, col_name in enumerate(columns, 1):
                    value = row[col_name]
                    # JSON-поля оставляем как строку (можно и распарсить, но для экспорта сойдёт)
                    ws.cell(row=row_idx, column=col_idx, value=value)

            # Автоширина столбцов (приблизительно)
            for col_idx, col_name in enumerate(columns, 1):
                max_length = len(str(col_name))
                for row in rows[:100]:  # ограничим для скорости
                    cell_value = str(row[col_name])
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    # Сохраняем во временный файл
    import tempfile
    import os
    fd, tmp_path = tempfile.mkstemp(suffix='.xlsx')
    os.close(fd)
    wb.save(tmp_path)

    filename = f"ma_furniture_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        tmp_path,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

# ========== РЕЖИМ ОБСЛУЖИВАНИЯ ==========
MAINTENANCE_FLAG=os.path.join(DATA_DIR,'.maintenance')
def is_maintenance_mode(): return os.path.exists(MAINTENANCE_FLAG)

@app.before_request
def handle_maintenance():
    exempt_paths=['/static/','/uploads/','/admin/','/api/admin/','/api/health','/favicon.ico','/api/admin/maintenance/status']
    if is_maintenance_mode():
        for path in exempt_paths:
            if request.path.startswith(path): return None
        maintenance_page=os.path.join(STATIC_DIR,'maintenance.html')
        if os.path.exists(maintenance_page): return send_file(maintenance_page),503
        else: return "<h1>Технические работы</h1><p>Сайт временно недоступен.</p>",503

@app.route('/api/admin/maintenance/status',methods=['GET'])
def admin_maintenance_status():
    try:
        enabled=is_maintenance_mode()
        return jsonify({'success':True,'enabled':enabled})
    except Exception as e: return jsonify({'success':False,'error':str(e)}),500

@app.route('/api/admin/maintenance/enable',methods=['POST'])
def admin_maintenance_enable():
    try:
        token=request.headers.get('Authorization')
        if not token: return jsonify({'success':False,'error':'Требуется авторизация'}),401
        with open(MAINTENANCE_FLAG,'w') as f: f.write(datetime.now().isoformat())
        logger.info("Режим технического обслуживания ВКЛЮЧЁН")
        return jsonify({'success':True,'message':'Режим обслуживания включён. Сайт недоступен для посетителей.'})
    except Exception as e: return jsonify({'success':False,'error':str(e)}),500

@app.route('/api/admin/maintenance/disable',methods=['POST'])
def admin_maintenance_disable():
    try:
        token=request.headers.get('Authorization')
        if not token: return jsonify({'success':False,'error':'Требуется авторизация'}),401
        if os.path.exists(MAINTENANCE_FLAG): os.remove(MAINTENANCE_FLAG)
        logger.info("Режим технического обслуживания ВЫКЛЮЧЁН")
        return jsonify({'success':True,'message':'Режим обслуживания выключен. Сайт снова доступен.'})
    except Exception as e: return jsonify({'success':False,'error':str(e)}),500

@app.route('/static/<path:filename>')
def serve_static(filename):
    return send_from_directory(STATIC_DIR, filename)

@app.route('/uploads/products/<filename>')
def serve_uploaded_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)

# 404 обработчик
@app.errorhandler(404)
def not_found_error(error):
    logger.warning(f"404 Not Found: {request.path}")
    if request.path == '/favicon.ico':
        return '', 204
    if request.path.startswith('/api/'):
        return jsonify({'success': False, 'error': 'Ресурс не найден'}), 404
    try:
        index_path = os.path.join(STATIC_DIR, 'index.html')
        if os.path.exists(index_path):
            return send_file(index_path)
        return jsonify({'success': False, 'error': 'Ресурс не найден'}), 404
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 404

if __name__=='__main__':
    print("\n"+"="*60)
    print("Starting Flask development server...")
    print("="*60)
    app.run(host='0.0.0.0',port=5000,debug=True)
else:
    print("\n"+"="*60)
    print("MA Furniture File Storage initialized")
    print("="*60)